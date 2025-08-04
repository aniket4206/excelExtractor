import pandas as pd
import numpy as np
from pathlib import Path
import argparse
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelComparator:
    def __init__(self, base_file_path):
        """
        Initialize the Excel Comparator with a base employee file
        
        Args:
            base_file_path (str): Path to the base Excel file containing employee data
        """
        self.base_file_path = base_file_path
        self.base_data = None
        self.comparison_results = {}
        
    def load_base_file(self):
        """Load the base Excel file"""
        try:
            self.base_data = pd.read_excel(self.base_file_path)
            print(f"✓ Base file loaded successfully: {len(self.base_data)} records")
            print(f"✓ Columns in base file: {list(self.base_data.columns)}")
            return True
        except Exception as e:
            print(f"✗ Error loading base file: {e}")
            return False
    
    def clean_data(self, df):
        """Clean and standardize data for comparison"""
        df_clean = df.copy()
        
        # Convert all columns to string and strip whitespace
        for col in df_clean.columns:
            if df_clean[col].dtype == 'object':
                df_clean[col] = df_clean[col].astype(str).str.strip().str.upper()
        
        return df_clean
    
    def find_common_columns(self, df1, df2):
        """Find common columns between two dataframes"""
        cols1 = set(df1.columns)
        cols2 = set(df2.columns)
        common_cols = list(cols1.intersection(cols2))
        return common_cols
    
    def compare_files(self, comparison_files, match_columns=None):
        """
        Compare multiple files against the base file
        
        Args:
            comparison_files (list): List of file paths to compare
            match_columns (list): Specific columns to use for matching (optional)
        """
        if self.base_data is None:
            print("✗ Please load the base file first")
            return
        
        base_clean = self.clean_data(self.base_data)
        
        for file_path in comparison_files:
            try:
                print(f"\n🔍 Processing: {file_path}")
                
                # Load comparison file
                comp_data = pd.read_excel(file_path)
                comp_clean = self.clean_data(comp_data)
                
                print(f"✓ Loaded {len(comp_data)} records from {Path(file_path).name}")
                
                # Find common columns
                common_cols = self.find_common_columns(base_clean, comp_clean)
                if not common_cols:
                    print(f"✗ No common columns found with base file")
                    continue
                
                # Use specified match columns or all common columns
                if match_columns:
                    match_cols = [col for col in match_columns if col in common_cols]
                    if not match_cols:
                        print(f"✗ None of the specified match columns found")
                        continue
                else:
                    match_cols = common_cols
                
                print(f"✓ Using columns for matching: {match_cols}")
                
                # Perform comparison
                result = self.perform_comparison(
                    base_clean, comp_clean, match_cols, Path(file_path).name
                )
                
                self.comparison_results[Path(file_path).name] = result
                
            except Exception as e:
                print(f"✗ Error processing {file_path}: {e}")
    
    def perform_comparison(self, base_df, comp_df, match_cols, file_name):
        """Perform detailed comparison between base and comparison dataframes"""
        
        # Create a composite key for matching
        if len(match_cols) == 1:
            base_df['_match_key'] = base_df[match_cols[0]]
            comp_df['_match_key'] = comp_df[match_cols[0]]
        else:
            base_df['_match_key'] = base_df[match_cols].apply(
                lambda x: '|'.join(x.astype(str)), axis=1
            )
            comp_df['_match_key'] = comp_df[match_cols].apply(
                lambda x: '|'.join(x.astype(str)), axis=1
            )
        
        # Find matches and misses
        base_keys = set(base_df['_match_key'].dropna())
        comp_keys = set(comp_df['_match_key'].dropna())
        
        matched_keys = base_keys.intersection(comp_keys)
        missing_in_comp = base_keys - comp_keys
        extra_in_comp = comp_keys - base_keys
        
        # Create result dataframes
        matched_base = base_df[base_df['_match_key'].isin(matched_keys)].copy()
        matched_comp = comp_df[comp_df['_match_key'].isin(matched_keys)].copy()
        missing_records = base_df[base_df['_match_key'].isin(missing_in_comp)].copy()
        extra_records = comp_df[comp_df['_match_key'].isin(extra_in_comp)].copy()
        
        # Remove the temporary match key
        for df in [matched_base, matched_comp, missing_records, extra_records]:
            if '_match_key' in df.columns:
                df.drop('_match_key', axis=1, inplace=True)
        
        result = {
            'file_name': file_name,
            'match_columns': match_cols,
            'total_base_records': len(base_df),
            'total_comp_records': len(comp_df),
            'matched_records': len(matched_keys),
            'missing_in_comparison': len(missing_in_comp),
            'extra_in_comparison': len(extra_in_comp),
            'matched_data_base': matched_base,
            'matched_data_comp': matched_comp,
            'missing_records': missing_records,
            'extra_records': extra_records
        }
        
        # Print summary
        print(f"📊 Comparison Summary for {file_name}:")
        print(f"   • Total records in base: {result['total_base_records']}")
        print(f"   • Total records in comparison: {result['total_comp_records']}")
        print(f"   • Matched records: {result['matched_records']}")
        print(f"   • Missing in comparison: {result['missing_in_comparison']}")
        print(f"   • Extra in comparison: {result['extra_in_comparison']}")
        
        return result
    
    def export_results(self, output_path="comparison_results.xlsx"):
        """Export comparison results to Excel with multiple sheets"""
        
        if not self.comparison_results:
            print("✗ No comparison results to export")
            return
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # Create summary sheet
            summary_data = []
            for file_name, result in self.comparison_results.items():
                summary_data.append({
                    'File Name': result['file_name'],
                    'Match Columns': ', '.join(result['match_columns']),
                    'Base Records': result['total_base_records'],
                    'Comparison Records': result['total_comp_records'],
                    'Matched': result['matched_records'],
                    'Missing in Comparison': result['missing_in_comparison'],
                    'Extra in Comparison': result['extra_in_comparison'],
                    'Match Rate %': round((result['matched_records'] / result['total_base_records']) * 100, 2) if result['total_base_records'] > 0 else 0
                })
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Create detailed sheets for each comparison
            for file_name, result in self.comparison_results.items():
                safe_name = file_name.replace('.xlsx', '').replace('.xls', '')[:31]  # Excel sheet name limit
                
                # Matched records sheet
                if not result['matched_data_base'].empty:
                    result['matched_data_base'].to_excel(
                        writer, sheet_name=f'{safe_name}_Matched', index=False
                    )
                
                # Missing records sheet
                if not result['missing_records'].empty:
                    result['missing_records'].to_excel(
                        writer, sheet_name=f'{safe_name}_Missing', index=False
                    )
                
                # Extra records sheet
                if not result['extra_records'].empty:
                    result['extra_records'].to_excel(
                        writer, sheet_name=f'{safe_name}_Extra', index=False
                    )
        
        # Apply formatting
        self.format_excel_output(output_path)
        print(f"✓ Results exported to: {output_path}")
    
    def format_excel_output(self, file_path):
        """Apply formatting to the Excel output"""
        try:
            wb = openpyxl.load_workbook(file_path)
            
            # Format summary sheet
            if 'Summary' in wb.sheetnames:
                ws = wb['Summary']
                
                # Header formatting
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            
        except Exception as e:
            print(f"⚠ Warning: Could not apply formatting: {e}")
    
    def print_detailed_summary(self):
        """Print a detailed summary of all comparisons"""
        print("\n" + "="*60)
        print("📋 DETAILED COMPARISON SUMMARY")
        print("="*60)
        
        for file_name, result in self.comparison_results.items():
            print(f"\n📁 {result['file_name']}")
            print("-" * 40)
            print(f"Match Columns: {', '.join(result['match_columns'])}")
            print(f"Base Records: {result['total_base_records']}")
            print(f"Comparison Records: {result['total_comp_records']}")
            print(f"✅ Matched: {result['matched_records']}")
            print(f"❌ Missing: {result['missing_in_comparison']}")
            print(f"➕ Extra: {result['extra_in_comparison']}")
            
            if result['total_base_records'] > 0:
                match_rate = (result['matched_records'] / result['total_base_records']) * 100
                print(f"📊 Match Rate: {match_rate:.2f}%")


def main():
    """Main function to run the Excel comparison tool"""
    parser = argparse.ArgumentParser(description='Compare Excel files against a base employee file')
    parser.add_argument('base_file', help='Path to the base Excel file')
    parser.add_argument('comparison_files', nargs='+', help='Paths to files to compare against base')
    parser.add_argument('--match-columns', nargs='+', help='Specific columns to use for matching')
    parser.add_argument('--output', default='comparison_results.xlsx', help='Output file path')
    
    args = parser.parse_args()
    
    print("🚀 Excel File Comparison Tool")
    print("="*50)
    
    # Initialize comparator
    comparator = ExcelComparator(args.base_file)
    
    # Load base file
    if not comparator.load_base_file():
        sys.exit(1)
    
    # Perform comparisons
    comparator.compare_files(args.comparison_files, args.match_columns)
    
    # Export results
    comparator.export_results(args.output)
    
    # Print summary
    comparator.print_detailed_summary()
    
    print(f"\n✅ Process completed! Check '{args.output}' for detailed results.")


if __name__ == "__main__":
    # Example usage when running directly
    if len(sys.argv) == 1:
        print("📖 Excel File Comparison Tool - Example Usage")
        print("="*50)
        print("Command line usage:")
        print("python excel_comparator.py base_file.xlsx file1.xlsx file2.xlsx file3.xlsx")
        print("\nWith specific match columns:")
        print("python excel_comparator.py base_file.xlsx file1.xlsx --match-columns EmployeeID Name")
        print("\nWith custom output:")
        print("python excel_comparator.py base_file.xlsx file1.xlsx --output my_results.xlsx")
        print("\n" + "="*50)
        print("📋 Interactive Mode:")
        
        # Interactive mode
        base_file = input("Enter path to base Excel file: ").strip()
        if not Path(base_file).exists():
            print("❌ Base file not found!")
            sys.exit(1)
        
        comparison_files = []
        print("\nEnter comparison file paths (press Enter when done):")
        while True:
            file_path = input("File path: ").strip()
            if not file_path:
                break
            if Path(file_path).exists():
                comparison_files.append(file_path)
                print(f"✅ Added: {file_path}")
            else:
                print(f"❌ File not found: {file_path}")
        
        if not comparison_files:
            print("❌ No comparison files provided!")
            sys.exit(1)
        
        # Ask for match columns
        match_cols_input = input("\nEnter specific columns for matching (comma-separated, or press Enter for auto-detect): ").strip()
        match_columns = [col.strip() for col in match_cols_input.split(',')] if match_cols_input else None
        
        output_file = input("Enter output file name (or press Enter for 'comparison_results.xlsx'): ").strip()
        if not output_file:
            output_file = 'comparison_results.xlsx'
        
        # Run comparison
        comparator = ExcelComparator(base_file)
        if comparator.load_base_file():
            comparator.compare_files(comparison_files, match_columns)
            comparator.export_results(output_file)
            comparator.print_detailed_summary()
            print(f"\n✅ Process completed! Check '{output_file}' for detailed results.")
    else:
        main()